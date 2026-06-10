import json
import os
import re
import warnings
from typing import Optional
from datetime import date, datetime, timedelta

try:
    from urllib3.exceptions import NotOpenSSLWarning
    warnings.filterwarnings('ignore', category=NotOpenSSLWarning)
except ImportError:
    pass

import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROVIDED_JIRA_BASE = "https://invitahealthtech.atlassian.net"
FALLBACK_JIRA_BASE = "https://invitahealth.atlassian.net"
BASE_CANDIDATES = [PROVIDED_JIRA_BASE, FALLBACK_JIRA_BASE]

GREEN = "70AD47"
RED = "FF0000"
ORANGE = "ED7D31"
LIGHT_BLUE = "BDD7EE"
YELLOW = "FFD966"
WHITE = "FFFFFF"

FILL_GREEN = PatternFill("solid", fgColor=GREEN)
FILL_RED = PatternFill("solid", fgColor=RED)
FILL_ORANGE = PatternFill("solid", fgColor=ORANGE)
FILL_LIGHT_BLUE = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_YELLOW = PatternFill("solid", fgColor=YELLOW)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

THIN = Side(style="thin", color="D9D9D9")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HYPERLINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
WRAP_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

DONE_STATUSES = {
    "Done", "Closed", "Resolved", "Released to Client", "Ready for Acceptance",
    "Not a Defect", "Duplicate", "Dev Approved",
}
NOT_STARTED_STATUSES = {"Not Started", "Open", "To Do", "Backlog"}
CR_PLUS_STATUSES = {
    "In Code Review", "Code Review", "Done", "Closed", "Resolved", "QA", "QA Testing",
    "QA Review", "Dev Approved", "Ready for Testing", "Ready For QA", "Testing",
    "In Testing", "Blocked (QA)", "Ready for Acceptance", "Released to Client",
    "Not a Defect", "Duplicate",
}


def ensure_parent(path):
    os.makedirs(os.path.dirname(path), exist_ok=True)


def save_json(path, data):
    ensure_parent(path)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)


def parse_jira_datetime(value):
    if not value:
        return None
    value = re.sub(r"([+-]\d{2})(\d{2})$", r"\1:\2", value)
    return datetime.fromisoformat(value)


def parse_iso_date(value):
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    return datetime.strptime(value[:10], "%Y-%m-%d").date()


def iso_date(value):
    dt = parse_jira_datetime(value) if isinstance(value, str) and "T" in value else parse_iso_date(value)
    return dt.date().isoformat() if isinstance(dt, datetime) else dt.isoformat() if dt else ""


def fmt_short_date(value):
    d = parse_iso_date(value)
    return d.strftime("%Y-%m-%d") if d else ""


def short_label(value):
    d = parse_iso_date(value)
    return d.strftime("%b %d, %Y") if d else ""


def days_between(start_value, end_value):
    start_dt = parse_jira_datetime(start_value) if isinstance(start_value, str) and "T" in start_value else None
    end_dt = parse_jira_datetime(end_value) if isinstance(end_value, str) and "T" in end_value else None
    if not start_dt:
        start_d = parse_iso_date(start_value)
        start_dt = datetime.combine(start_d, datetime.min.time()) if start_d else None
    if not end_dt:
        end_d = parse_iso_date(end_value)
        end_dt = datetime.combine(end_d, datetime.min.time()) if end_d else None
    if not start_dt or not end_dt:
        return None
    return round((end_dt - start_dt).total_seconds() / 86400.0, 1)


def pct(value, total):
    return (value / total) if total else 0


def pct_text(value, total, decimals=0):
    return f"{pct(value, total):.{decimals}%}" if total else "0%"


def safe_float(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def first_value(*values):
    for value in values:
        if value not in (None, ""):
            return value
    return None


def story_points(fields):
    return first_value(fields.get("customfield_10132"), fields.get("customfield_10016"), fields.get("customfield_10096"))


def latest_sprint_name(sprints):
    if not sprints:
        return "No Sprint"
    latest = max(sprints, key=lambda sprint: int(sprint.get("id", 0) or 0))
    return latest.get("name") or "No Sprint"


def sprint_scope_name(sprint_name, target="TK Sprint 26-2"):
    return "In Scope" if sprint_name == target else "Out of Scope"


def jira_assignee_name(fields):
    assignee = fields.get("assignee") or {}
    return assignee.get("displayName") or "Unassigned"


def jira_priority_name(fields):
    priority = fields.get("priority") or {}
    return priority.get("name") or ""


def jira_status_name(fields):
    status = fields.get("status") or {}
    return status.get("name") or ""


def jira_issue_type(fields):
    issue_type = fields.get("issuetype") or {}
    return issue_type.get("name") or ""


def dev_type_summary(records):
    counts = {}
    for record in records:
        counts[record["type"]] = counts.get(record["type"], 0) + 1
    return ", ".join(f"{name}: {counts[name]}" for name in sorted(counts)) if counts else "—"


def week_windows(today=None):
    today = today or date.today()
    start = today - timedelta(days=27)
    windows = []
    for index in range(4):
        win_start = start + timedelta(days=index * 7)
        win_end = win_start + timedelta(days=6)
        windows.append((win_start, win_end))
    return windows


def week_label(start, end):
    return f"{start.strftime('%b %d')} – {end.strftime('%b %d')}"


def current_week_start(today=None):
    today = today or date.today()
    return today - timedelta(days=6)


class JiraClient:
    def __init__(self):
        self.user = os.environ.get("JIRA_USER")
        self.token = os.environ.get("JIRA_TOKEN")
        if not self.user or not self.token:
            raise RuntimeError("JIRA_USER and JIRA_TOKEN must be set in the environment.")
        self.session = requests.Session()
        self.session.auth = (self.user, self.token)
        self.session.headers.update({"Accept": "application/json", "Content-Type": "application/json"})
        self.base_url = None

    def _request(self, method, path, **kwargs):
        last_error = None
        for base in BASE_CANDIDATES:
            url = f"{base}{path}"
            response = self.session.request(method, url, timeout=120, **kwargs)
            if response.ok:
                self.base_url = base
                return response
            last_error = f"{base} -> {response.status_code}: {response.text[:200]}"
            if response.status_code not in {401, 404}:
                response.raise_for_status()
        raise RuntimeError(last_error or f"Jira request failed for {path}")

    def search_jql(self, jql, fields, expand=None):
        issues = []
        payload = {"jql": jql, "maxResults": 100, "fields": fields}
        if expand:
            payload["expand"] = expand
        while True:
            response = self._request("POST", "/rest/api/3/search/jql", json=payload)
            data = response.json()
            batch = data.get("issues", [])
            issues.extend(batch)
            next_token = data.get("nextPageToken")
            if not next_token:
                break
            payload["nextPageToken"] = next_token
        return issues

    def get_issue(self, key, fields=None):
        params = {}
        if fields:
            params["fields"] = ",".join(fields)
        response = self._request("GET", f"/rest/api/3/issue/{key}", params=params)
        return response.json()

    def get_user_groups(self, account_id):
        """Return list of Jira group names for a given account ID."""
        resp = self._request("GET", f"/rest/api/3/user?accountId={account_id}&expand=groups")
        return [g["name"] for g in resp.json().get("groups", {}).get("items", [])]

    def get_issue_changelog(self, key):
        histories = []
        start_at = 0
        while True:
            response = self._request(
                "GET",
                f"/rest/api/3/issue/{key}/changelog",
                params={"startAt": start_at, "maxResults": 100},
            )
            data = response.json()
            values = data.get("values") or data.get("histories") or []
            histories.extend(values)
            max_results = data.get("maxResults", len(values) or 100)
            total = data.get("total", len(histories))
            start_at += max_results
            if start_at >= total or not values:
                break
        return histories


# ---------------------------------------------------------------------------
# User Role Classification
# ---------------------------------------------------------------------------

# Maps Jira group name substrings to canonical department labels.
# Evaluated in order — first match wins.
_GROUP_TO_DEPT = [
    ("DNA&F - Developers",  "Software Engineering"),
    ("DNA&F - QE",          "Quality Engineering"),
    ("DNA&F - FAS",         "FAS"),
]

# Known overrides for accounts whose Jira groups don't reflect their actual team.
# Format: { account_id: department }
_DEPT_OVERRIDES = {
    "712020:eced4f7e-81c5-4d4a-94a3-91e6dfc35feb": "Software Engineering",  # Diyaa Ebrahim (contractor)
    "712020:f98bc3e5-1d52-44e7-ae46-9ec44a274de4": "Quality Engineering",   # Donald Davis (no QE group assigned)
}


class UserRoleClassifier:
    """Classifies Jira users into departments based on group membership.

    Results are cached for the lifetime of the instance so bulk classification
    across many tickets doesn't hammer the API.

    Usage:
        classifier = UserRoleClassifier(jira_client)
        dept = classifier.get_department(account_id)
        # Returns one of: "Software Engineering", "Quality Engineering", "FAS", or None
    """

    def __init__(self, jira: "JiraClient"):
        self._jira = jira
        self.        _cache: dict = {}

    def get_department(self, account_id: str) -> Optional[str]:
        """Return the department for an account_id, or None if unclassifiable."""
        if not account_id:
            return None
        if account_id in self._cache:
            return self._cache[account_id]

        # Check hard overrides first
        if account_id in _DEPT_OVERRIDES:
            dept = _DEPT_OVERRIDES[account_id]
            self._cache[account_id] = dept
            return dept

        try:
            groups = self._jira.get_user_groups(account_id)
        except Exception:
            self._cache[account_id] = None
            return None

        dept = None
        for group_fragment, label in _GROUP_TO_DEPT:
            if any(group_fragment in g for g in groups):
                dept = label
                break

        self._cache[account_id] = dept
        return dept

    def annotate_issues(self, issues: list[dict], fields: list[str] = None) -> list[dict]:
        """Add '_reporter_dept' and '_assignee_dept' to each issue dict in-place.

        Args:
            issues: List of Jira issue dicts (from search_jql).
            fields:  Which user fields to classify. Defaults to ['reporter', 'assignee'].

        Returns:
            The same list with dept keys added.
        """
        if fields is None:
            fields = ["reporter", "assignee"]
        for issue in issues:
            f = issue.get("fields", {})
            for field in fields:
                user = f.get(field)
                aid = user.get("accountId") if user else None
                issue[f"_{field}_dept"] = self.get_department(aid)
        return issues
